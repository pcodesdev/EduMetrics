"""
report_builder.py — PDF and Excel report generation.

Generates:
- School Report PDF  (cover, summary, charts, tables, at-risk, recommendations)
- Class Report PDF   (class vs school comparison, color-coded student table)
- Student Report PDF (scores, trend chart, rank, risk, remarks)
- Excel Export       (cleaned data, computed columns, conditional formatting, per-class sheets)

All PDFs are A4, print-ready with school name / date footer.
"""

import io
import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import matplotlib
matplotlib.use("Agg")  # Non-interactive backend for server use
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, inch, mm
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from core.stats import compute_overview, compute_subject_stats, compute_student_profile
from core.risk import compute_risk_scores
from core.kenya_grading import get_grade_label, classify_system
from core.insights import generate_all_insights


# ── Colour palette ──────────────────────────────────────────────────

BRAND_DARK  = colors.HexColor("#1a1a2e")
BRAND_MID   = colors.HexColor("#16213e")
BRAND_ACCENT = colors.HexColor("#0f3460")
BRAND_HIGHLIGHT = colors.HexColor("#e94560")
GREEN       = colors.HexColor("#2ecc71")
AMBER       = colors.HexColor("#f39c12")
RED         = colors.HexColor("#e74c3c")
LIGHT_GREY  = colors.HexColor("#f5f5f5")
WHITE       = colors.white

MPL_PALETTE = ["#0f3460", "#e94560", "#2ecc71", "#f39c12", "#9b59b6", "#1abc9c"]


# ── Helpers ─────────────────────────────────────────────────────────

def _safe_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _find_col(df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
    cols_lower = {c.lower().strip(): c for c in df.columns}
    for alias in aliases:
        if alias.lower() in cols_lower:
            return cols_lower[alias.lower()]
    return None


def _footer(canvas, doc, school_name: str):
    """Draw school name and date in the page footer."""
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.grey)
    footer_text = f"{school_name} — Generated {datetime.now().strftime('%d %B %Y, %H:%M')}"
    canvas.drawString(2 * cm, 1.2 * cm, footer_text)
    canvas.drawRightString(A4[0] - 2 * cm, 1.2 * cm, f"Page {doc.page}")
    canvas.restoreState()


def _chart_to_image(fig, width=14 * cm, height=8 * cm) -> Image:
    """Convert a matplotlib figure to a ReportLab Image."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return Image(buf, width=width, height=height)


def _ensure_percentage(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure a percentage column exists."""
    df = df.copy()
    if "percentage" not in df.columns:
        score_col = _find_col(df, ["score", "marks", "mark", "total_score"])
        max_col = _find_col(df, ["max_score", "max_marks", "out_of", "maximum"])
        if score_col and max_col:
            s = pd.to_numeric(df[score_col], errors="coerce")
            m = pd.to_numeric(df[max_col], errors="coerce")
            df["percentage"] = (s / m * 100).round(2)
        elif score_col:
            df["percentage"] = pd.to_numeric(df[score_col], errors="coerce")
    return df


def _split_class_stream(class_value: Any) -> (str, str):
    """
    Split class string into class level and stream where possible.
    Example: 'Form 2A' -> ('Form 2', 'A'), fallback stream='N/A'.
    """
    cls = str(class_value or "N/A").strip()
    if cls == "N/A":
        return "N/A", "N/A"
    parts = cls.split()
    if not parts:
        return cls, "N/A"
    tail = parts[-1]
    if len(tail) >= 2 and tail[-1].isalpha() and any(ch.isdigit() for ch in tail):
        return f"{' '.join(parts[:-1])} {tail[:-1]}".strip(), tail[-1].upper()
    if len(parts) >= 2 and parts[-1].isalpha() and len(parts[-1]) <= 2:
        return " ".join(parts[:-1]).strip(), parts[-1].upper()
    if len(cls) >= 2 and cls[-1].isalpha() and any(ch.isdigit() for ch in cls):
        return cls[:-1].strip(), cls[-1].upper()
    return cls, "N/A"


def _draw_security_marks(canvas, verification_code: str):
    """Draw faint anti-forgery marks/signature text in the background."""
    canvas.saveState()
    canvas.setFillColor(colors.Color(0.75, 0.75, 0.75, alpha=0.16))
    canvas.setFont("Helvetica-Bold", 34)
    canvas.translate(4.5 * cm, 13.5 * cm)
    canvas.rotate(32)
    canvas.drawString(0, 0, "OFFICIAL SCHOOL REPORT")
    canvas.restoreState()


def _detect_grading_system(df: pd.DataFrame, class_hint: Optional[str] = None) -> str:
    """Infer grading system from class label; defaults to secondary."""
    class_name = class_hint
    if not class_name:
        class_col = _find_col(df, ["class", "grade", "form", "stream"])
        if class_col and not df.empty:
            class_name = str(df.iloc[0].get(class_col, ""))
    return classify_system(str(class_name or ""))

    # Embedded repeated verification stamp text.
    canvas.saveState()
    canvas.setFillColor(colors.Color(0.65, 0.65, 0.65, alpha=0.14))
    canvas.setFont("Helvetica-Oblique", 8)
    y = 3.2 * cm
    for _ in range(7):
        canvas.drawString(2.0 * cm, y, f"Verified Signature: Principal | Code: {verification_code}")
        y += 3.0 * cm
    canvas.restoreState()


# ── Charts ──────────────────────────────────────────────────────────

def _subject_bar_chart(subject_stats: Dict[str, Any]) -> Image:
    """Bar chart of mean score per subject."""
    subjects_data = subject_stats.get("subjects", [])
    if not subjects_data:
        return None

    names = [s.get("subject", "?") for s in subjects_data]
    means = [_safe_float(s.get("mean", 0)) or 0 for s in subjects_data]

    fig, ax = plt.subplots(figsize=(8, 4))
    bars = ax.bar(names, means, color=MPL_PALETTE[:len(names)], edgecolor="white", linewidth=0.5)

    for bar, val in zip(bars, means):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1,
                f"{val:.1f}", ha="center", va="bottom", fontsize=8, fontweight="bold")

    ax.set_ylabel("Mean Score (%)", fontsize=10)
    ax.set_title("Subject Performance Overview", fontsize=12, fontweight="bold", pad=12)
    ax.set_ylim(0, 105)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(axis="x", rotation=30)
    fig.tight_layout()
    return _chart_to_image(fig)


def _score_distribution_chart(overview: Dict[str, Any]) -> Image:
    """Histogram-style chart of score distribution."""
    # Backend returns: distribution: {bins: [...], counts: [...]}
    dist = overview.get("distribution", {})
    if not dist or not dist.get("bins"):
        return None

    labels = dist["bins"]
    values = dist["counts"]

    fig, ax = plt.subplots(figsize=(7, 3.5))
    bar_colors = ["#e74c3c", "#f39c12", "#f1c40f", "#2ecc71", "#27ae60"]
    bar_colors = bar_colors[:len(labels)]
    ax.bar(labels, values, color=bar_colors, edgecolor="white", linewidth=0.5)
    ax.set_ylabel("Number of Students", fontsize=10)
    ax.set_title("Score Distribution", fontsize=12, fontweight="bold", pad=12)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    return _chart_to_image(fig, width=13 * cm, height=7 * cm)


def _student_trend_chart(profile: Dict[str, Any]) -> Image:
    """Line chart showing a student's score trend across terms."""
    trends = profile.get("term_trends", [])
    if not trends:
        return None

    terms = [t.get("term", "?") for t in trends]
    means = [_safe_float(t.get("mean", 0)) or 0 for t in trends]

    fig, ax = plt.subplots(figsize=(6, 3))
    ax.plot(terms, means, marker="o", color=MPL_PALETTE[0], linewidth=2, markersize=8)
    ax.fill_between(range(len(terms)), means, alpha=0.15, color=MPL_PALETTE[0])

    for i, (t, m) in enumerate(zip(terms, means)):
        ax.annotate(f"{m:.1f}", (i, m), textcoords="offset points",
                    xytext=(0, 10), ha="center", fontsize=8, fontweight="bold")

    ax.set_ylabel("Mean Score (%)", fontsize=10)
    ax.set_title("Performance Trend", fontsize=12, fontweight="bold", pad=12)
    ax.set_ylim(0, 105)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    return _chart_to_image(fig, width=12 * cm, height=6 * cm)


def _student_radar_chart(profile: Dict[str, Any]) -> Image:
    """Radar chart of a student's latest subject scores."""
    subject_scores = profile.get("subject_scores", [])
    if not subject_scores:
        return None

    labels = [s.get("subject", "?") for s in subject_scores]
    # Backend returns {subject, score} — no separate latest_score field
    values = [_safe_float(s.get("score", 0)) or 0 for s in subject_scores]

    num = len(labels)
    if num < 3:
        return None

    angles = np.linspace(0, 2 * np.pi, num, endpoint=False).tolist()
    values_closed = values + values[:1]
    angles_closed = angles + angles[:1]

    fig, ax = plt.subplots(figsize=(5, 5), subplot_kw=dict(polar=True))
    ax.fill(angles_closed, values_closed, color=MPL_PALETTE[0], alpha=0.2)
    ax.plot(angles_closed, values_closed, color=MPL_PALETTE[0], linewidth=2)
    ax.set_xticks(angles)
    ax.set_xticklabels(labels, fontsize=8)
    ax.set_ylim(0, 100)
    ax.set_title("Subject Scores", fontsize=12, fontweight="bold", pad=20)
    fig.tight_layout()
    return _chart_to_image(fig, width=10 * cm, height=10 * cm)


def _student_subject_bar_chart(profile: Dict[str, Any], pass_mark: int) -> Optional[Image]:
    """Simple subject bar chart for parent-friendly reading."""
    subject_scores = profile.get("subject_scores", [])
    if not subject_scores:
        return None

    subjects = [str(s.get("subject", "?")) for s in subject_scores]
    values = [_safe_float(s.get("score", 0)) or 0 for s in subject_scores]
    colors_list = [
        "#2ecc71" if v >= 70 else "#f39c12" if v >= pass_mark else "#e74c3c"
        for v in values
    ]

    fig, ax = plt.subplots(figsize=(7.2, 3.6))
    bars = ax.bar(subjects, values, color=colors_list, edgecolor="white", linewidth=0.8)
    for bar, val in zip(bars, values):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 1.5,
            f"{val:.0f}%",
            ha="center",
            va="bottom",
            fontsize=8,
            fontweight="bold",
        )

    ax.axhline(pass_mark, linestyle="--", color="#7f8c8d", linewidth=1)
    ax.text(len(subjects) - 0.5, pass_mark + 1, f"Pass mark {pass_mark}%", fontsize=8, color="#7f8c8d")
    ax.set_ylim(0, 100)
    ax.set_xlabel("Subjects", fontsize=9)
    ax.set_ylabel("Score (%)", fontsize=9)
    ax.set_title("Subject Scores At a Glance", fontsize=11, fontweight="bold")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(axis="x", rotation=28, labelsize=8)
    fig.tight_layout()
    return _chart_to_image(fig, width=15 * cm, height=7.5 * cm)


def _term_trend_chart_from_overview(overview: Dict[str, Any]) -> Optional[Image]:
    term_trends = overview.get("term_trends", [])
    if not term_trends:
        return None

    terms = [str(t.get("term", "?")) for t in term_trends]
    means = [_safe_float(t.get("mean", 0)) or 0 for t in term_trends]
    fig, ax = plt.subplots(figsize=(7, 3.5))
    ax.plot(terms, means, marker="o", linewidth=2.2, color="#0f3460")
    ax.fill_between(range(len(terms)), means, alpha=0.15, color="#0f3460")
    for i, m in enumerate(means):
        ax.text(i, m + 1.2, f"{m:.1f}%", ha="center", fontsize=8)
    ax.set_ylim(0, 100)
    ax.set_ylabel("Mean Score (%)")
    ax.set_xlabel("Term")
    ax.set_title("Term-by-Term Trend")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    return _chart_to_image(fig, width=14 * cm, height=6.5 * cm)


def _pass_fail_donut_chart(overview: Dict[str, Any]) -> Optional[Image]:
    p = int(overview.get("pass_count", 0) or 0)
    f = int(overview.get("fail_count", 0) or 0)
    if (p + f) <= 0:
        return None
    fig, ax = plt.subplots(figsize=(4.5, 4.0))
    wedges, _ = ax.pie([p, f], colors=["#2ecc71", "#e74c3c"], startangle=90, wedgeprops={"width": 0.42})
    ax.legend(wedges, [f"Pass ({p})", f"Fail ({f})"], loc="lower center", bbox_to_anchor=(0.5, -0.08), ncol=2, fontsize=8)
    ax.set_title("Pass/Fail Split")
    fig.tight_layout()
    return _chart_to_image(fig, width=7 * cm, height=6.5 * cm)


def _risk_level_bar_chart(risk_data: Dict[str, Any]) -> Optional[Image]:
    summary = risk_data.get("summary", {})
    vals = [
        int(summary.get("high_risk", 0) or 0),
        int(summary.get("medium_risk", 0) or 0),
        int(summary.get("low_risk", 0) or 0),
    ]
    if sum(vals) == 0:
        return None
    labels = ["High", "Medium", "Low"]
    cols = ["#e74c3c", "#f39c12", "#2ecc71"]
    fig, ax = plt.subplots(figsize=(5.8, 3.5))
    bars = ax.bar(labels, vals, color=cols)
    for b, v in zip(bars, vals):
        ax.text(b.get_x() + b.get_width() / 2, b.get_height() + 0.2, str(v), ha="center", fontsize=8)
    ax.set_title("Risk Levels")
    ax.set_ylabel("Students")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    return _chart_to_image(fig, width=9 * cm, height=6.5 * cm)


# ── PDF Helpers ─────────────────────────────────────────────────────

def _styles():
    """Return custom paragraph styles."""
    ss = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle(
            "CustomTitle", parent=ss["Title"],
            fontSize=28, leading=34, textColor=BRAND_DARK,
            spaceAfter=6 * mm,
        ),
        "subtitle": ParagraphStyle(
            "CustomSubtitle", parent=ss["Normal"],
            fontSize=14, leading=18, textColor=BRAND_ACCENT,
            spaceAfter=4 * mm,
        ),
        "heading": ParagraphStyle(
            "CustomHeading", parent=ss["Heading2"],
            fontSize=14, leading=18, textColor=BRAND_DARK,
            spaceBefore=8 * mm, spaceAfter=4 * mm,
        ),
        "body": ParagraphStyle(
            "CustomBody", parent=ss["Normal"],
            fontSize=10, leading=14, textColor=colors.black,
            spaceAfter=3 * mm,
        ),
        "small": ParagraphStyle(
            "CustomSmall", parent=ss["Normal"],
            fontSize=8, leading=10, textColor=colors.grey,
        ),
        "center": ParagraphStyle(
            "CenterBody", parent=ss["Normal"],
            fontSize=10, leading=14, alignment=TA_CENTER,
        ),
    }
    return styles


def _make_table(data: List[List], col_widths=None, header_color=BRAND_DARK):
    """Create a styled table."""
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), header_color),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT_GREY]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle(style_cmds))
    return t


def _color_coded_table(data: List[List], score_col_idx: int, pass_mark: int,
                       col_widths=None):
    """Table with per-row colour coding based on a score column."""
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]

    for row_idx in range(1, len(data)):
        try:
            score = float(data[row_idx][score_col_idx])
        except (ValueError, TypeError, IndexError):
            continue
        if score >= 70:
            bg = colors.HexColor("#d5f5e3")
        elif score >= pass_mark:
            bg = colors.HexColor("#fef9e7")
        else:
            bg = colors.HexColor("#fadbd8")
        style_cmds.append(("BACKGROUND", (0, row_idx), (-1, row_idx), bg))

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle(style_cmds))
    return t


# ═══════════════════════════════════════════════════════════════════
# 1. SCHOOL REPORT PDF
# ═══════════════════════════════════════════════════════════════════

def generate_school_report_pdf(
    output_path: str,
    school_name: str,
    overview: Dict[str, Any],
    subject_stats: Dict[str, Any],
    risk_data: Dict[str, Any],
    gap_data: Dict[str, Any],
    insights: Dict[str, Any],
    pass_mark: int = 50,
    school_system: str = "secondary",
):
    """Generate a comprehensive school performance report PDF."""
    st = _styles()
    story = []

    ov = overview
    school_grade = get_grade_label(_safe_float(ov.get("overall_mean")), school_system)
    subjects = subject_stats.get("subjects", [])
    at_risk = risk_data.get("students", risk_data.get("at_risk_students", []))

    # ── Page 1: Executive cover ────────────────────────────────────
    story.append(Spacer(1, 5.5 * cm))
    story.append(Paragraph(school_name, st["title"]))
    story.append(Paragraph("Comprehensive School Performance Report", st["subtitle"]))
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(datetime.now().strftime("%d %B %Y"), st["body"]))
    story.append(Spacer(1, 1.5 * cm))
    story.append(Paragraph(
        "This report highlights overall performance, hidden trends, risk patterns, "
        "and actionable recommendations for leadership and teaching teams.",
        st["body"],
    ))
    exec_summary = insights.get("executive_summary", "")
    if exec_summary:
        story.append(Spacer(1, 5 * mm))
        story.append(Paragraph(f"<b>Executive Summary:</b> {exec_summary}", st["body"]))
    story.append(PageBreak())

    # ── Page 2: Core metrics + visuals ─────────────────────────────
    story.append(Paragraph("1) School Performance Snapshot", st["heading"]))
    overview_data = [
        ["Metric", "Value"],
        ["Total Students", str(ov.get("total_students", "N/A"))],
        ["Total Records", str(ov.get("total_records", "N/A"))],
        ["Overall Mean", f"{_safe_float(ov.get('overall_mean', 0)) or 0:.1f}%"],
        ["Overall Grade", school_grade],
        ["Median", f"{_safe_float(ov.get('overall_median', 0)) or 0:.1f}%"],
        ["Std Deviation", f"{_safe_float(ov.get('overall_std', 0)) or 0:.1f}"],
        ["Pass Rate", f"{_safe_float(ov.get('pass_rate', 0)) or 0:.1f}%"],
        ["Fail Rate", f"{_safe_float(ov.get('fail_rate', 0)) or 0:.1f}%"],
        ["Pass Mark", f"{pass_mark}%"],
    ]
    story.append(_make_table(overview_data, col_widths=[7.5 * cm, 6.5 * cm]))
    story.append(Spacer(1, 5 * mm))

    term_chart = _term_trend_chart_from_overview(overview)
    if term_chart:
        story.append(Paragraph("2) Trend Across Terms", st["heading"]))
        story.append(term_chart)
        story.append(Spacer(1, 4 * mm))

    subj_chart = _subject_bar_chart(subject_stats)
    if subj_chart:
        story.append(Paragraph("3) Subject Mean Performance", st["heading"]))
        story.append(subj_chart)
    story.append(PageBreak())

    # ── Page 3: Distribution, risk, deep details ───────────────────
    story.append(Paragraph("4) Distribution and Hidden Performance Details", st["heading"]))
    dist_chart = _score_distribution_chart(overview)
    if dist_chart:
        story.append(dist_chart)
        story.append(Spacer(1, 4 * mm))

    donut = _pass_fail_donut_chart(overview)
    risk_bar = _risk_level_bar_chart(risk_data)
    if donut or risk_bar:
        row = []
        widths = []
        if donut:
            row.append(donut)
            widths.append(7.2 * cm)
        if risk_bar:
            row.append(risk_bar)
            widths.append(7.2 * cm)
        visual_tbl = Table([row], colWidths=widths)
        visual_tbl.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        story.append(visual_tbl)
        story.append(Spacer(1, 5 * mm))

    if subjects:
        story.append(Paragraph("5) Subject Detail Table", st["heading"]))
        subj_table_data = [["Subject", "Mean", "Median", "Std", "Pass %", "Fail %"]]
        for s in subjects:
            subj_table_data.append([
                str(s.get("subject", "?")),
                f"{_safe_float(s.get('mean', 0)) or 0:.1f}",
                f"{_safe_float(s.get('median', 0)) or 0:.1f}",
                f"{_safe_float(s.get('std', s.get('std_dev', 0)) or 0):.1f}",
                f"{_safe_float(s.get('pass_rate', 0)) or 0:.1f}%",
                f"{_safe_float(s.get('fail_rate', 0)) or 0:.1f}%",
            ])
        story.append(_make_table(subj_table_data))
        story.append(Spacer(1, 4 * mm))

    if at_risk:
        story.append(Paragraph("6) At-Risk Students (Top 20 by risk score)", st["heading"]))
        risk_table = [["Name", "ID", "Class", "Avg %", "Risk", "Level"]]
        for s in at_risk[:20]:
            risk_table.append([
                str(s.get("name", "?")),
                str(s.get("student_id", "?")),
                str(s.get("class", "?")),
                f"{_safe_float(s.get('overall_mean', 0)) or 0:.1f}",
                f"{_safe_float(s.get('risk_score', 0)) or 0:.0f}",
                str(s.get("risk_level", "?")),
            ])
        story.append(_make_table(risk_table))
        story.append(Spacer(1, 4 * mm))

    all_insights = insights.get("insights", [])
    if all_insights:
        story.append(Paragraph("7) Key Recommendations", st["heading"]))
        for ins in all_insights[:10]:
            title = str(ins.get("title", "Insight"))
            narrative = str(ins.get("narrative", ins.get("description", "")))
            story.append(Paragraph(f"<b>{title}</b>: {narrative}", st["small"]))

    # ── Build PDF ───────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2.5 * cm,
    )
    doc.build(
        story,
        onFirstPage=lambda c, d: _footer(c, d, school_name),
        onLaterPages=lambda c, d: _footer(c, d, school_name),
    )


# ═══════════════════════════════════════════════════════════════════
# 2. CLASS REPORT PDF
# ═══════════════════════════════════════════════════════════════════

def generate_class_report_pdf(
    output_path: str,
    school_name: str,
    class_name: str,
    class_df: pd.DataFrame,
    school_df: pd.DataFrame,
    pass_mark: int = 50,
):
    """Generate a class-level performance report PDF."""
    st = _styles()
    story = []

    class_df = _ensure_percentage(class_df)
    school_df = _ensure_percentage(school_df)

    # ── Derived aggregates ──────────────────────────────────────────
    school_mean = school_df["percentage"].mean() if "percentage" in school_df.columns else 0
    class_mean = class_df["percentage"].mean() if "percentage" in class_df.columns else 0
    id_col = _find_col(class_df, ["student_id", "studentid", "id", "adm_no"])
    name_col = _find_col(class_df, ["name", "student_name", "full_name"])
    subj_col = _find_col(class_df, ["subject", "course", "paper"])
    term_col = _find_col(class_df, ["term", "semester"])

    class_students = class_df[id_col].nunique() if id_col else len(class_df)
    school_students = school_df[id_col].nunique() if id_col and id_col in school_df.columns else len(school_df)
    class_pass_rate = (class_df["percentage"] >= pass_mark).mean() * 100 if "percentage" in class_df.columns else 0
    school_pass_rate = (school_df["percentage"] >= pass_mark).mean() * 100 if "percentage" in school_df.columns else 0
    grading_system = _detect_grading_system(class_df, class_name)
    class_grade = get_grade_label(class_mean, grading_system)
    school_grade = get_grade_label(school_mean, grading_system)

    # ── Page 1: Cover + KPI ─────────────────────────────────────────
    story.append(Spacer(1, 4.5 * cm))
    story.append(Paragraph(school_name, st["title"]))
    story.append(Paragraph(f"Class Performance Report — {class_name}", st["subtitle"]))
    story.append(Paragraph(datetime.now().strftime("%d %B %Y"), st["body"]))
    story.append(Spacer(1, 8 * mm))
    comp_data = [
        ["Metric", "Class", "School"],
        ["Students", str(class_students), str(school_students)],
        ["Mean Score", f"{class_mean:.1f}%", f"{school_mean:.1f}%"],
        ["Grade", class_grade, school_grade],
        ["Pass Rate", f"{class_pass_rate:.1f}%", f"{school_pass_rate:.1f}%"],
    ]
    story.append(_make_table(comp_data, col_widths=[6 * cm, 4 * cm, 4 * cm]))
    story.append(Spacer(1, 6 * mm))

    ov_class = compute_overview(class_df, pass_mark=pass_mark)
    term_chart = _term_trend_chart_from_overview(ov_class)
    if term_chart:
        story.append(Paragraph("Class Trend by Term", st["heading"]))
        story.append(term_chart)
    story.append(PageBreak())

    # ── Page 2: Subject and distribution deep dive ─────────────────
    story.append(Paragraph("Subject Deep Dive", st["heading"]))
    if subj_col and "percentage" in class_df.columns:
        subj_table = [["Subject", "Class Mean", "School Mean", "Class Pass %", "Gap"]]
        for subj in sorted(class_df[subj_col].dropna().unique()):
            c_scores = class_df[class_df[subj_col] == subj]["percentage"]
            s_scores = school_df[school_df[subj_col] == subj]["percentage"] if subj_col in school_df.columns else pd.Series()
            c_mean = c_scores.mean() if len(c_scores) else np.nan
            s_mean = s_scores.mean() if len(s_scores) else np.nan
            gap = c_mean - s_mean if pd.notna(c_mean) and pd.notna(s_mean) else np.nan
            subj_table.append([
                str(subj),
                f"{c_mean:.1f}" if pd.notna(c_mean) else "N/A",
                f"{s_mean:.1f}" if pd.notna(s_mean) else "N/A",
                f"{(c_scores >= pass_mark).mean() * 100:.1f}%" if len(c_scores) else "N/A",
                f"{gap:+.1f}" if pd.notna(gap) else "N/A",
            ])
        story.append(_make_table(subj_table))
        story.append(Spacer(1, 5 * mm))

    cls_subject_stats = compute_subject_stats(class_df, pass_mark=pass_mark)
    cls_subj_chart = _subject_bar_chart(cls_subject_stats)
    if cls_subj_chart:
        story.append(cls_subj_chart)
        story.append(Spacer(1, 5 * mm))

    class_dist = _score_distribution_chart(ov_class)
    if class_dist:
        story.append(Paragraph("Class Score Distribution", st["heading"]))
        story.append(class_dist)
    story.append(PageBreak())

    # ── Page 3: Students, ranking, risk ────────────────────────────
    story.append(Paragraph("Student-Level Details", st["heading"]))
    if id_col and "percentage" in class_df.columns:
        student_avgs = class_df.groupby(id_col).agg(
            name=(name_col, "first") if name_col else (id_col, "first"),
            avg_pct=("percentage", "mean"),
        ).reset_index().sort_values("avg_pct", ascending=False)

        student_table = [["Rank", "Student", "Avg %", "Status"]]
        for rank, (_, row) in enumerate(student_avgs.iterrows(), 1):
            avg = row["avg_pct"]
            status = "Pass" if avg >= pass_mark else "Fail"
            student_table.append([str(rank), str(row.get("name", row[id_col])), f"{avg:.1f}", status])
        story.append(_color_coded_table(student_table, score_col_idx=2, pass_mark=pass_mark))
        story.append(Spacer(1, 4 * mm))

    risk_data = compute_risk_scores(class_df, pass_mark=pass_mark)
    at_risk = risk_data.get("students", risk_data.get("at_risk_students", []))
    risk_chart = _risk_level_bar_chart(risk_data)
    if risk_chart:
        story.append(risk_chart)
        story.append(Spacer(1, 3 * mm))

    if at_risk:
        story.append(Paragraph("At-Risk Students in Class", st["heading"]))
        risk_tbl = [["Name", "Risk Score", "Level", "Average %"]]
        for s in at_risk[:15]:
            risk_tbl.append([
                str(s.get("name", "?")),
                f"{_safe_float(s.get('risk_score', 0)) or 0:.0f}",
                str(s.get("risk_level", "?")),
                f"{_safe_float(s.get('overall_mean', 0)) or 0:.1f}",
            ])
        story.append(_make_table(risk_tbl, col_widths=[6 * cm, 2.5 * cm, 2.5 * cm, 3 * cm]))
        story.append(Spacer(1, 4 * mm))

    # ── Class-specific insights ────────────────────────────────────
    class_insights = generate_all_insights(class_df, pass_mark=pass_mark)
    insight_items = class_insights.get("insights", []) if isinstance(class_insights, dict) else []
    student_id_col = _find_col(class_df, ["student_id", "studentid", "id", "adm_no", "admission_no"])
    student_name_col = _find_col(class_df, ["name", "student_name", "full_name", "student"])
    name_to_id = {}
    id_to_name = {}
    if student_id_col and student_name_col:
        for _, r in class_df[[student_name_col, student_id_col]].dropna().iterrows():
            nm = str(r.get(student_name_col, "")).strip()
            sid = str(r.get(student_id_col, "")).strip()
            if nm and sid:
                name_to_id[nm.lower()] = sid
                id_to_name[sid.lower()] = nm
    if insight_items:
        story.append(Paragraph("Key Insights for This Class", st["heading"]))
        for ins in insight_items[:10]:
            title = str(ins.get("title", "Insight"))
            narrative = str(ins.get("narrative", ins.get("description", "")))
            recommendation = str(ins.get("recommendation", "")).strip()
            severity = str(ins.get("severity", "info")).upper()
            supporting = ins.get("supporting_data", {}) if isinstance(ins.get("supporting_data"), dict) else {}
            insight_student_name = str(
                supporting.get("student")
                or supporting.get("name")
                or ""
            ).strip()
            insight_student_id = str(
                supporting.get("student_id")
                or supporting.get("admission_no")
                or ""
            ).strip()
            if insight_student_name and not insight_student_id:
                insight_student_id = name_to_id.get(insight_student_name.lower(), "")
            if insight_student_id and not insight_student_name:
                insight_student_name = id_to_name.get(insight_student_id.lower(), "")
            # If insight "name" is actually an admission number, resolve the true name.
            if insight_student_name and not insight_student_id:
                inferred_name = id_to_name.get(insight_student_name.lower())
                if inferred_name:
                    insight_student_id = insight_student_name
                    insight_student_name = inferred_name
            story.append(Paragraph(f"<b>[{severity}] {title}</b>: {narrative}", st["small"]))
            if insight_student_name:
                sid_text = insight_student_id if insight_student_id else "N/A"
                story.append(Paragraph(f"Student: <b>{insight_student_name}</b> | Admission No: <b>{sid_text}</b>", st["small"]))
            if recommendation:
                story.append(Paragraph(f"Action: {recommendation}", st["small"]))
            story.append(Spacer(1, 1.5 * mm))
    else:
        story.append(Paragraph("Key Insights for This Class", st["heading"]))
        story.append(Paragraph("No significant class-specific insights were generated for this dataset.", st["small"]))

    # ── Build ───────────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2.5 * cm,
    )
    doc.build(
        story,
        onFirstPage=lambda c, d: _footer(c, d, school_name),
        onLaterPages=lambda c, d: _footer(c, d, school_name),
    )


# ═══════════════════════════════════════════════════════════════════
# 3. STUDENT REPORT CARD PDF
# ═══════════════════════════════════════════════════════════════════

def generate_student_report_pdf(
    output_path: str,
    school_name: str,
    student_id: str,
    df: pd.DataFrame,
    pass_mark: int = 50,
):
    """Generate a parent-friendly 2-page student report card PDF."""
    st = _styles()
    story = []

    source_df = _ensure_percentage(df.copy())
    profile = compute_student_profile(source_df, student_id, pass_mark=pass_mark)
    if not profile:
        # Minimal fallback
        story.append(Paragraph(f"No data found for student {student_id}.", st["body"]))
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        doc.build(story)
        return

    name = str(profile.get("name", student_id))
    cls_raw = str(profile.get("class", "N/A"))
    class_level, class_stream = _split_class_stream(cls_raw)
    school_value = str(profile.get("school", school_name))
    overall_mean = _safe_float(profile.get("overall_mean", 0)) or 0.0
    class_rank = profile.get("class_rank", "N/A")
    class_total = profile.get("class_total", "N/A")
    grading_system = _detect_grading_system(source_df, cls_raw)
    student_grade = get_grade_label(overall_mean, grading_system)

    class_mean = None
    school_mean = None
    class_grade = "—"
    overall_grade = "—"
    class_col = _find_col(source_df, ["class", "grade", "form", "stream"])
    if "percentage" in source_df.columns:
        school_mean = _safe_float(source_df["percentage"].mean())
    if class_col and cls_raw and "percentage" in source_df.columns:
        class_subset = source_df[
            source_df[class_col].astype(str).str.strip().str.lower() == str(cls_raw).strip().lower()
        ]
        if not class_subset.empty:
            class_mean = _safe_float(class_subset["percentage"].mean())
    class_grade = get_grade_label(class_mean, grading_system)
    overall_grade = get_grade_label(school_mean, grading_system)

    subject_scores = profile.get("subject_scores", [])
    best_subject = "N/A"
    needs_support = "N/A"
    if subject_scores:
        sorted_subj = sorted(
            [(str(s.get("subject", "?")), _safe_float(s.get("score", 0)) or 0.0) for s in subject_scores],
            key=lambda x: x[1],
            reverse=True,
        )
        best_subject = f"{sorted_subj[0][0]} ({sorted_subj[0][1]:.0f}%)"
        needs_support = f"{sorted_subj[-1][0]} ({sorted_subj[-1][1]:.0f}%)"

    # Teacher remarks and recommendations (simple rule-based narrative)
    if overall_mean >= 75:
        teacher_remark = "Excellent and consistent effort. The learner demonstrates strong mastery across most subjects."
    elif overall_mean >= pass_mark:
        teacher_remark = "Good progress. The learner meets expectations and should continue steady revision."
    else:
        teacher_remark = "Performance is below target. The learner needs close support and a structured improvement plan."

    teacher_recommendations = [
        f"1. Prioritize weekly revision in the weakest subject: {needs_support}.",
        "2. Maintain a daily study timetable and complete all assignments on time.",
        "3. Attend remedial/consultation sessions and track progress after each assessment.",
    ]

    # ── PAGE 1: Identity + simple summary + visual ─────────────────
    story.append(Paragraph(school_value, st["title"]))
    story.append(Paragraph("Student Report Card (Parent-Friendly)", st["subtitle"]))
    story.append(Paragraph(datetime.now().strftime("%d %B %Y"), st["small"]))
    story.append(Spacer(1, 4 * mm))

    identity = [
        ["Student Name", name, "Admission Number", str(student_id)],
        ["Class", class_level, "Stream", class_stream],
        ["School", school_value, "Academic Year", str(datetime.now().year)],
    ]
    identity_table = Table(identity, colWidths=[3.8 * cm, 3.7 * cm, 3.8 * cm, 3.7 * cm])
    identity_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#d1d5db")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(identity_table)
    story.append(Spacer(1, 5 * mm))

    parent_summary = [
        ["Overall Score", f"{overall_mean:.1f}%"],
        ["Student Grade", student_grade],
        ["Class Grade", class_grade],
        ["Overall Grade", overall_grade],
        ["Class Position", f"{class_rank} of {class_total}"],
        ["Best Subject", best_subject],
        ["Needs More Support", needs_support],
    ]
    summary_table = Table(parent_summary, colWidths=[6.5 * cm, 8.5 * cm])
    summary_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef2ff")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 4 * mm))

    simple_remark = (
        "How to read this report: Green means strong performance, amber means improving, and red means more support is needed. "
        f"The target is at least {pass_mark}% in each subject."
    )
    story.append(Paragraph(simple_remark, st["body"]))
    story.append(Spacer(1, 2 * mm))

    subject_chart = _student_subject_bar_chart(profile, pass_mark=pass_mark)
    if subject_chart:
        story.append(subject_chart)
    story.append(PageBreak())

    # ── PAGE 2: Trend + detailed table + signatures ────────────────
    story.append(Paragraph("Progress Over Time", st["heading"]))
    trend_img = _student_trend_chart(profile)
    if trend_img:
        story.append(trend_img)
    else:
        story.append(Paragraph("No term trend data available yet.", st["body"]))
    story.append(Spacer(1, 3 * mm))

    story.append(Paragraph("Subject-by-Subject Breakdown", st["heading"]))
    if subject_scores:
        scores_data = [["Subject", "Average Score", "Meaning"]]
        max_rows = 12
        for s in subject_scores[:max_rows]:
            score = _safe_float(s.get("score", 0)) or 0
            meaning = "Doing Well" if score >= 70 else "Keep Improving" if score >= pass_mark else "Needs Support"
            scores_data.append([s.get("subject", "?"), f"{score:.1f}%", meaning])
        story.append(
            _color_coded_table(
                scores_data,
                score_col_idx=1,
                pass_mark=pass_mark,
                col_widths=[7.5 * cm, 3.5 * cm, 4 * cm],
            )
        )
        if len(subject_scores) > max_rows:
            story.append(Paragraph(
                f"Note: Showing first {max_rows} subjects to keep this report within 2 pages.",
                st["small"],
            ))
    else:
        story.append(Paragraph("No subject score data available.", st["body"]))
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph("Home Support Tips", st["heading"]))
    tips = [
        "1. Ask your child to revise daily for 30-45 minutes.",
        "2. Check homework completion every evening.",
        "3. Encourage extra practice in the weakest subject.",
    ]
    for tip in tips:
        story.append(Paragraph(tip, st["body"]))
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("Class Teacher's Remarks & Recommendations", st["heading"]))
    story.append(Paragraph(teacher_remark, st["body"]))
    for rec in teacher_recommendations:
        story.append(Paragraph(rec, st["body"]))
    story.append(Spacer(1, 4 * mm))

    verification_code = f"{student_id}-{datetime.now().strftime('%Y%m%d')}"
    story.append(Paragraph(f"Verification Code: <b>{verification_code}</b>", st["small"]))
    story.append(Spacer(1, 3 * mm))

    sign_table = Table(
        [
            ["Class Teacher Signature", "Parent/Guardian Signature", "Principal Signature"],
            ["", "", ""],
            ["Date: ____________", "Date: ____________", "Date: ____________"],
        ],
        colWidths=[5 * cm, 5 * cm, 5 * cm],
        rowHeights=[0.65 * cm, 1.5 * cm, 0.55 * cm],
    )
    sign_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9ca3af")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
    ]))
    story.append(sign_table)

    # ── Build ───────────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2.5 * cm,
    )

    def _student_page_decor(c, d):
        _draw_security_marks(c, verification_code)
        _footer(c, d, school_name)

    doc.build(
        story,
        onFirstPage=_student_page_decor,
        onLaterPages=_student_page_decor,
    )


# ═══════════════════════════════════════════════════════════════════
# 4. EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════

def generate_excel_export(
    output_path: str,
    df: pd.DataFrame,
    school_name: str,
    pass_mark: int = 50,
):
    """Export cleaned data to Excel with computed columns, conditional formatting, per-class sheets."""
    df = _ensure_percentage(df)
    df = df.copy()

    # Compute helper columns
    id_col = _find_col(df, ["student_id", "studentid", "id", "adm_no"])
    name_col = _find_col(df, ["name", "student_name", "full_name"])
    class_col = _find_col(df, ["class", "grade", "stream", "form"])

    # Add average column per student
    if id_col and "percentage" in df.columns:
        avgs = df.groupby(id_col)["percentage"].mean().rename("student_avg")
        df = df.merge(avgs, on=id_col, how="left")

        # Add rank
        if id_col:
            rank_map = avgs.rank(ascending=False, method="min").astype(int).rename("rank")
            df = df.merge(rank_map, on=id_col, how="left")

    # Styling definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    red_fill = PatternFill(start_color="fadbd8", end_color="fadbd8", fill_type="solid")
    green_fill = PatternFill(start_color="d5f5e3", end_color="d5f5e3", fill_type="solid")
    yellow_fill = PatternFill(start_color="fef9e7", end_color="fef9e7", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _style_sheet(ws, dataframe):
        """Apply formatting to a worksheet."""
        # Header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Find percentage column index
        pct_col_idx = None
        for idx, col_name in enumerate(dataframe.columns, 1):
            if col_name == "percentage":
                pct_col_idx = idx
                break

        # Apply conditional formatting and borders
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            if pct_col_idx and row[pct_col_idx - 1].value is not None:
                try:
                    val = float(row[pct_col_idx - 1].value)
                    fill = green_fill if val >= 70 else (yellow_fill if val >= pass_mark else red_fill)
                    for cell in row:
                        cell.fill = fill
                except (ValueError, TypeError):
                    pass

        # Freeze header
        ws.freeze_panes = "A2"

        # Auto-width columns
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 30)

    wb = Workbook()

    # ── Sheet 1: All Students ───────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Students"
    ws_all.sheet_properties.tabColor = "1a1a2e"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws_all.append(row)
    _style_sheet(ws_all, df)

    # ── Per-class sheets ────────────────────────────────────────────
    if class_col:
        classes = sorted(df[class_col].dropna().unique())
        tab_colors = ["0f3460", "e94560", "2ecc71", "f39c12", "9b59b6", "1abc9c"]
        for i, cls in enumerate(classes):
            cls_df = df[df[class_col] == cls].copy()
            safe_name = str(cls).replace("/", "-")[:28]
            ws = wb.create_sheet(title=safe_name)
            ws.sheet_properties.tabColor = tab_colors[i % len(tab_colors)]
            for row in dataframe_to_rows(cls_df, index=False, header=True):
                ws.append(row)
            _style_sheet(ws, cls_df)

    wb.save(output_path)
